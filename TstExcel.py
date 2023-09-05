# Importando módulos necessários
import inspect
import re
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
import funcsexcel as fxml  # Importa um módulo funcsexcel, presumivelmente personalizado
import pandas as pd
import numpy as np
import os

# Obtenha o diretório atual do seu arquivo .py
diretorio_atual = os.path.dirname(__file__)

def letras_para_numeros(texto):
    texto = texto.lower()
    valor_total = 0
    
    for letra in texto:
        if 'a' <= letra <= 'z':
            valor_total = valor_total * 26 + (ord(letra) - ord('a') + 1)
    
    return valor_total - 1

# Função que identifica se um valor é um número
def identifica_numero(coluna):
    for valor in coluna:
        if valor == np.NaN:
            return False
        try:
            float(valor)
            return True
        except:
            return False

# Definindo uma função para listar os argumentos necessários para chamar uma função
def listar_argumentos(funcao):
    
    assinatura = inspect.signature(funcao)
    parametros = assinatura.parameters
    argumentos_necessarios = []

    # Iterando pelos parâmetros da função
    for parametro, info in parametros.items():
        if info.default == inspect.Parameter.empty:
            argumentos_necessarios.append(parametro)

    return argumentos_necessarios

# Definindo uma função para verificar se uma coluna tem uma fórmula
def has_formula(x):
    with open(os.path.join(diretorio_atual, '..', '1_ENTRADA', 'Teste.txt', 'r') as file:
        v = False
        for line in file:
            s = f'Coluna {x} '
            if s in line:
                v = True
        return v

# Definindo uma função para extrair linhas de um arquivo e criar um dicionário com os dados
def extract_lines(tdict):
    extracting = False

    with open(os.path.join(diretorio_atual, '..', '1_ENTRADA', 'Teste.txt', 'r') as file:
        for line in file:
            if extracting:
                if '------------------------------' in line:
                    break
                l = line.split(': ')
                tdict[l[0]] = l[1].replace('\n','')
            elif 'Dicionario de dados' in line:
                extracting = True

# Criando um dicionário para armazenar dados extraídos
dict_dados = {}
extract_lines(dict_dados)

# Criando um dicionário para armazenar informações de funções
dict_func = {}

# Iterando pelos dados extraídos
for i in dict_dados:
    if has_formula(i):
        stri = 'fxml.calcular_'+i

        # Listando os argumentos necessários para a função
        dfi = listar_argumentos(eval(stri))
        atts = []

        # Criando uma lista de argumentos formatados
        for a in dfi:
            if a.isalpha():
                a = a.lower()
            atts.append(f'df_aux[dict_dados[\'{a}\']]')

        dict_func[i] = atts

# Caminho do arquivo Excel de entrada
path = os.path.join(diretorio_atual, '..', '1_ENTRADA', 'Teste_real.xlsx'
a = pd.read_excel(path, header=1)
df = pd.DataFrame(a)
df_aux = df.copy()

xl = pd.ExcelFile(path)
nomes_planilhas = xl.sheet_names

d_sheets = {}
for i in range(1, len(nomes_planilhas)+1):
    k = f'S{i}'
    n_pls = nomes_planilhas[i-1]
    a_df = pd.read_excel(xl, sheet_name=n_pls)
    out_df = pd.DataFrame(a_df)
    d_sheets[k] = a_df

workbook = openpyxl.Workbook()
sheet = workbook.active
sheet_input = workbook.create_sheet(title="Input")

# Definir o estilo da fonte para a primeira linha (caracteres brancos)
font = Font(color="FFFFFF")  # Branco

# Definir o estilo das bordas para a primeira linha (bordas pretas)
border = Border(left=Side(border_style="thin", color="000000"), 
                right=Side(border_style="thin", color="000000"),
                top=Side(border_style="thin", color="000000"),
                bottom=Side(border_style="thin", color="000000"))

# Abrindo um arquivo de saída para escrever resultados
with open(diretorio_atual, '..', '2_SAIDA', "RelAngelo1.txt", 'w') as f:

    # Iterando pelos dados extraídos
    for x,y in dict_dados.items():
        col_want = y
        
        # Se a coluna não tiver uma fórmula, continue para a próxima
        if x not in dict_func:
            data = df_aux[col_want].tolist()

            # Adicionar os dados à primeira coluna não preenchida
            column = 1
            while sheet_input.cell(row=1, column=column).value is not None:
                column += 1

            # Adicionar o cabeçalho da coluna
            sheet_input.cell(row=1, column=column, value=col_want)

            for i, value in enumerate(data, start=1):
                sheet_input.cell(row=i+1, column=column, value=value)  # +1 para evitar a sobreposição com o cabeçalho

            continue

        try:
            # Construindo uma string de função e avaliando-a
            correspondencias = re.findall('S\d_\w_\d', ' '.join(dict_func[x]))

            attr_str = ', '.join(map(str, dict_func[x]))
            for cor in list(set(correspondencias)):
                var_sub = f'df_aux[dict_dados[\'{cor}\']]'
                if var_sub in attr_str:
                    sht = cor.split("_")[0]
                    cell = cor.split("_")[1]
                    pos = int(cor.split("_")[2])
                    nan = "NaN"
                    vals_cor = list(d_sheets[sht][d_sheets[sht].columns[letras_para_numeros(cell)]])[pos]
                    new_cor = f'{vals_cor}'
                    attr_str = attr_str.replace(var_sub, new_cor)

            func_str = f'np.vectorize(fxml.calcular_{x})({attr_str})'
            print(func_str)
            df_aux[col_want] = eval(func_str)
        except Exception as exception:
            print(str(exception))
            f.write(x + ' : ' + 'algo errado (exception/error)\n')
            continue

        # Verificando se as colunas são iguais ou têm valores nulos
        if df_aux[col_want].astype(str).equals(df[col_want].astype(str)):
            f.write(x + ' : ' + 'tudo certo\n')
        elif identifica_numero(df_aux[col_want]) and identifica_numero(df[col_want]):
            if df_aux[col_want].astype(int, errors='ignore').equals(df[col_want].astype(int, errors='ignore')):
                f.write(x + ' : ' + 'tudo certo\n')
        elif (df[col_want].isnull().sum() + df_aux[col_want].isnull().sum()) == (len(df[col_want]) + len(df_aux[col_want])):
            f.write(x + ' : ' + 'tudo certo\n')
        else:
            f.write(x + ' : ' + 'algo errado\n')

        # Adicionar os dados à primeira coluna não preenchida
        sheet.title = 'Output'
        data = df_aux[col_want].tolist()

        # Adicionar os dados à primeira coluna não preenchida
        column = 1
        while sheet.cell(row=1, column=column).value is not None:
            column += 1

        # Adicionar o cabeçalho da coluna
        sheet.cell(row=1, column=column, value=col_want)

        for i, value in enumerate(data, start=1):
            sheet.cell(row=i+1, column=column, value=value)  # +1 para evitar a sobreposição com o cabeçalho

for cell in sheet[1]:
    cell.fill = PatternFill(start_color="FFD633", end_color="FFD633", fill_type="solid")  # amarelo
    cell.border = border
    cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

# Definir a cor de preenchimento para todas as células na primeira linha
for cell in sheet_input[1]:
    cell.fill = PatternFill(start_color="00517B", end_color="00517B", fill_type="solid")  # Azul
    cell.font = font
    cell.border = border
    cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

workbook.save(path = os.path.join(diretorio_atual, '..', '2_SAIDA', 'Teste_real_output.xlsx')